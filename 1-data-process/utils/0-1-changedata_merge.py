"""
# @Author: 算法组
# @Date: 2026-05-14
# @Description: 数据整理工具：
#   1. 从源文件夹按通道（VIS / WIDE / ZOOM）提取图像，复制到 OUTPUT/{通道}/网格名/日期/。
#   2. 筛选、重命名后，各通道分别拢并为 GEOAI-new-{通道}/A/B 结构。
# @Branch: 忽略单次飞行网格，同一天多次只保留一次架次文件夹。
# @Filter: 每个网格保留时间差最大的 2 个日期（配对样本数须一致），其余删除。
# @Rename: Train-{项目}-vis-08.01.JPG（纯 ASCII：RENAME_HEAD + 通道小写 + 网格编号 + 序号，无中文）。
# @Excel : 各通道分别生成统计 Excel。
# @GEOAI : 各通道拢并为 GEOAI-new-VIS / GEOAI-new-WIDE / GEOAI-new-ZOOM
# @Command: python 0-changedata_merge.py
"""

import os
import re
import shutil
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

#==============================#
# 接口配置
#==============================#
INPUT_DIR = r"D:\0-data\1-ChangeDetect\cache\数据中心文件-中科青云"
RENAME_HEAD = "Train-zhongke"       
OUTPUT_DIR = INPUT_DIR                
#==============================#

# 通道名 -> 文件名后缀（大写匹配）
IMAGE_CHANNELS = {
    "VIS": "_VIS.JPG",
    "WIDE": "_WIDE.JPG",
    "ZOOM": "_ZOOM.JPG",
}

# 网格文件夹名须包含该关键字（天池小学网格01、天池网格08 等均满足）
GRID_NAME_KEY = "网格"
#==============================#

SKIP_NAMES = {"黑龙镇.zip"}
CHANNEL_DIR_NAMES = set(IMAGE_CHANNELS.keys())
GEOAI_PREFIX = "GEOAI-new"


def _is_grid_folder(name: str) -> bool:
    """判断是否为网格目录（名称含「网格」）。"""
    return GRID_NAME_KEY in name


def _grid_folder_name(grid_name: str) -> str:
    """输出目录中的网格文件夹名：保留原名，避免「天池小学网格08」与「天池网格08」冲突。"""
    return grid_name


def _rename_head_base() -> str:
    """Train-luoboxiaoxue（去掉末尾连字符）。"""
    return RENAME_HEAD.rstrip("-_")


def _grid_file_id(grid_name: str) -> str:
    """
    重命名用网格编号（仅 ASCII）：取文件夹名「网格」后的编号部分。
    例：天池小学网格08 -> 08；网格007_018 -> 007-018
    """
    if GRID_NAME_KEY in grid_name:
        part = grid_name.split(GRID_NAME_KEY, 1)[1]
    else:
        part = grid_name
    part = part.strip("-_")
    if not part:
        digits = re.findall(r"\d+", grid_name)
        part = digits[-1] if digits else "00"
    if part.isdigit():
        return part.zfill(2) if len(part) <= 2 else part
    # 007_018 等：只保留数字与分隔符
    part = part.replace("_", "-")
    part = re.sub(r"[^0-9\-]", "", part) or "00"
    return part


def build_train_filename(channel: str, grid_name: str, seq_index: int) -> str:
    """
    生成训练文件名：Train-luoboxiaoxue-vis-08.01.JPG
    seq_index 从 0 起，输出序号 01、02...
    """
    base = _rename_head_base()
    ch = channel.lower()
    gid = _grid_file_id(grid_name)
    return f"{base}-{ch}-{gid}.{seq_index + 1:02d}.JPG"


def _channel_suffix(channel: str) -> str:
    """返回通道对应文件名后缀（大写）。"""
    return IMAGE_CHANNELS[channel].upper()


def _match_channel(filename: str) -> str | None:
    """根据文件名判断所属通道，不匹配返回 None。"""
    upper = filename.upper()
    for ch, suf in IMAGE_CHANNELS.items():
        if suf.upper() in upper:
            return ch
    return None


def collect_images(input_dir: str) -> list:
    """遍历输入目录，按通道收集图像信息。"""
    records = []

    for grid_name in sorted(os.listdir(input_dir)):
        grid_path = os.path.join(input_dir, grid_name)
        if not os.path.isdir(grid_path):
            continue
        if grid_name in CHANNEL_DIR_NAMES or grid_name.startswith(GEOAI_PREFIX):
            continue
        if not _is_grid_folder(grid_name):
            continue

        all_subdirs = sorted(
            d for d in os.listdir(grid_path)
            if os.path.isdir(os.path.join(grid_path, d))
        )

        #------------#
        # 忽略单次飞行
        #------------#
        if len(all_subdirs) <= 1:
            print(f"  [跳过] {grid_name} (仅 {len(all_subdirs)} 次飞行)")
            continue

        #------------#
        # 按日期分组
        #------------#
        date_folders = defaultdict(list)
        for sub in all_subdirs:
            m = re.search(r"(\d{4}-\d{2}-\d{2})", sub)
            if m:
                date_folders[m.group(1).replace("-", "")].append(sub)

        #------------#
        # 收集各通道图像
        #------------#
        for date_key in sorted(date_folders):
            chosen = sorted(date_folders[date_key])[0]
            sub_path = os.path.join(grid_path, chosen)

            for filename in os.listdir(sub_path):
                channel = _match_channel(filename)
                if channel is None:
                    continue
                records.append({
                    "channel": channel,
                    "grid": grid_name,
                    "date": date_key,
                    "src": os.path.join(sub_path, filename),
                    "name": filename,
                })

    return records


def copy_images(records: list, output_dir: str) -> dict:
    """复制到 OUTPUT/{通道}/网格名/日期/ 结构。"""
    totals = defaultdict(int)

    for rec in records:
        grid_folder = _grid_folder_name(rec["grid"])
        dst = os.path.join(
            output_dir, rec["channel"], grid_folder, rec["date"],
        )
        os.makedirs(dst, exist_ok=True)
        dst_file = os.path.join(dst, rec["name"])

        if not os.path.exists(dst_file):
            shutil.copy2(rec["src"], dst_file)
            totals[rec["channel"]] += 1

    for ch in IMAGE_CHANNELS:
        print(f"  [{ch}] 复制 {totals[ch]} 张")
    return dict(totals)


def _iter_grid_dirs(channel_dir: str):
    """遍历通道目录下的网格文件夹。"""
    if not os.path.isdir(channel_dir):
        return
    for grid in sorted(os.listdir(channel_dir)):
        gp = os.path.join(channel_dir, grid)
        if os.path.isdir(gp) and _is_grid_folder(grid):
            yield grid, gp


def select_two_dates(channel_dir: str) -> int:
    """每个网格保留样本数一致的两个日期，其余删除。返回删除的日期文件夹数。"""
    total_deleted = 0
    skipped = []

    for grid, gp in _iter_grid_dirs(channel_dir):
        date_info = []
        for d in sorted(os.listdir(gp)):
            dp = os.path.join(gp, d)
            if not os.path.isdir(dp):
                continue
            cnt = len([f for f in os.listdir(dp) if f.upper().endswith(".JPG")])
            if cnt > 0:
                date_info.append((d, cnt))

        if len(date_info) < 2:
            skipped.append((grid, "日期不足2个"))
            continue

        dates = date_info
        latest = dates[-1]
        chosen = None
        for i in range(len(dates) - 1):
            if dates[i][1] == latest[1]:
                chosen = (dates[i][0], latest[0])
                break

        if chosen is None:
            skipped.append((grid, [f"{d}({c})" for d, c in dates]))
            continue

        keep_dates = set(chosen)
        for d, _ in dates:
            if d not in keep_dates:
                shutil.rmtree(os.path.join(gp, d))
                total_deleted += 1

        print(f"  {grid:<24} 保留 {chosen[0]} ~ {chosen[1]}  ({dates[0][1]}张)")

    if skipped:
        for g, v in skipped:
            print(f"  [跳过] {g:<24} {v}")

    return total_deleted


def rename_images(channel_dir: str, channel: str) -> int:
    """各日期下该通道图像数量一致时，统一重命名。"""
    suffix = _channel_suffix(channel)
    renamed_total = 0
    skipped = []

    for grid, gp in _iter_grid_dirs(channel_dir):
        date_dirs = sorted(os.listdir(gp))

        date_files = {}
        for d in date_dirs:
            dp = os.path.join(gp, d)
            if not os.path.isdir(dp):
                continue
            files = sorted([f for f in os.listdir(dp) if suffix in f.upper()])
            if files:
                date_files[d] = files

        if not date_files:
            continue

        counts = [len(v) for v in date_files.values()]
        if len(set(counts)) != 1:
            skipped.append((grid, counts))
            continue

        sample_count = counts[0]
        for k in range(sample_count):
            new_name = build_train_filename(channel, grid, k)
            for date_dir in sorted(date_files):
                src_file = os.path.join(gp, date_dir, date_files[date_dir][k])
                dst_file = os.path.join(gp, date_dir, new_name)
                if os.path.normcase(src_file) == os.path.normcase(dst_file):
                    continue
                os.rename(src_file, dst_file)
                renamed_total += 1

        print(f"  {grid:<24} 已重命名 {sample_count} 张 x {len(date_files)} 天")

    if skipped:
        print("  [跳过] 样本数不一致的网格：")
        for g, cnts in skipped:
            print(f"    {g:<24} {cnts}")

    return renamed_total


def generate_excel(channel_dir: str, channel: str):
    """生成该通道的统计 Excel。"""
    data = []
    for grid, gp in _iter_grid_dirs(channel_dir):
        for date in sorted(os.listdir(gp)):
            dp = os.path.join(gp, date)
            if not os.path.isdir(dp):
                continue
            cnt = len([f for f in os.listdir(dp) if f.upper().endswith(".JPG")])
            data.append((grid, date, cnt))

    wb = Workbook()
    ws = wb.active
    ws.title = "统计"

    h_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill("solid", fgColor="4472C4")
    d_font = Font(name="Arial", size=10)
    d_align = Alignment(horizontal="center", vertical="center")
    fill_a = PatternFill("solid", fgColor="D9E2F3")
    fill_b = PatternFill("solid", fgColor="FFFFFF")
    b_thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for c, h in enumerate(["网格", "日期", "数量"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = d_align
        cell.border = b_thin

    r = 2
    i = 0
    while i < len(data):
        grid = data[i][0]
        j = i
        while j < len(data) and data[j][0] == grid:
            j += 1
        n = j - i

        for k in range(n):
            row = r + k
            fill = fill_a if row % 2 == 0 else fill_b
            for col, val in [(2, data[i + k][1]), (3, data[i + k][2])]:
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = d_font
                cell.alignment = d_align
                cell.border = b_thin
                cell.fill = fill

        if n > 0:
            ws.merge_cells(start_row=r, start_column=1, end_row=r + n - 1, end_column=1)
            cell = ws.cell(row=r, column=1, value=grid)
            cell.font = d_font
            cell.alignment = d_align
            cell.border = b_thin
            cell.fill = fill_a if r % 2 == 0 else fill_b

        r += n
        i = j

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.freeze_panes = "A2"

    excel_path = os.path.join(channel_dir, f"统计-{channel}.xlsx")
    wb.save(excel_path)
    print(f"  Excel 已保存：{excel_path}")


#-------------#
# GEOAI 拢并
#-------------#


def _organize_one_grid(
    grid_dir: str,
    root_a: str,
    root_b: str,
    root_json: str,
    root_label: str,
):
    """整理单个网格子文件夹：两期日期或 A/B → 母目录。"""
    name = os.path.basename(grid_dir)
    print(f"\n========== {name} ==========")
    items = sorted(os.listdir(grid_dir))

    date_dirs = [
        d for d in items
        if os.path.isdir(os.path.join(grid_dir, d))
        and d.isdigit() and len(d) == 8
    ]
    has_grid_ab = (
        os.path.isdir(os.path.join(grid_dir, "A"))
        and os.path.isdir(os.path.join(grid_dir, "B"))
    )

    if len(date_dirs) == 2:
        old_dir, new_dir = date_dirs
        for fname in os.listdir(os.path.join(grid_dir, old_dir)):
            shutil.move(
                os.path.join(grid_dir, old_dir, fname),
                os.path.join(root_a, fname),
            )
        os.rmdir(os.path.join(grid_dir, old_dir))
        print(f"  {old_dir}  → A/")
        for fname in os.listdir(os.path.join(grid_dir, new_dir)):
            shutil.move(
                os.path.join(grid_dir, new_dir, fname),
                os.path.join(root_b, fname),
            )
        os.rmdir(os.path.join(grid_dir, new_dir))
        print(f"  {new_dir}  → B/")

    elif has_grid_ab:
        for fname in os.listdir(os.path.join(grid_dir, "A")):
            shutil.move(
                os.path.join(grid_dir, "A", fname),
                os.path.join(root_a, fname),
            )
        os.rmdir(os.path.join(grid_dir, "A"))
        print("  网格内 A/ → A/")
        for fname in os.listdir(os.path.join(grid_dir, "B")):
            shutil.move(
                os.path.join(grid_dir, "B", fname),
                os.path.join(root_b, fname),
            )
        os.rmdir(os.path.join(grid_dir, "B"))
        print("  网格内 B/ → B/")
    else:
        print("  ⚠ 跳过: 未找到 2 个日期文件夹或 A/B")
        return

    for sub, root in (("json", root_json), ("label", root_label)):
        src = os.path.join(grid_dir, sub)
        if os.path.isdir(src):
            for fname in os.listdir(src):
                shutil.move(os.path.join(src, fname), os.path.join(root, fname))
            os.rmdir(src)
            print(f"  {sub}/ → {sub}/")


def organize_geoai(channel_dir: str, output_base: str, channel: str):
    """将单通道 网格/日期 拢并为 GEOAI-new-{通道}/A/B/json/label。"""
    out_dir = os.path.join(output_base, f"{GEOAI_PREFIX}-{channel}")
    out_name = os.path.basename(out_dir)
    root_a = os.path.join(out_dir, "A")
    root_b = os.path.join(out_dir, "B")
    root_json = os.path.join(out_dir, "json")
    root_label = os.path.join(out_dir, "label")
    for p in (root_a, root_b, root_json, root_label):
        os.makedirs(p, exist_ok=True)

    count = 0
    skip_top = SKIP_NAMES | CHANNEL_DIR_NAMES | {f"{GEOAI_PREFIX}-{c}" for c in IMAGE_CHANNELS}

    for entry in sorted(os.listdir(channel_dir)):
        if entry in skip_top or entry.startswith(GEOAI_PREFIX):
            continue
        if entry in ("A", "B", "json", "label"):
            continue
        grid_dir = os.path.join(channel_dir, entry)
        if not os.path.isdir(grid_dir):
            continue
        if not _is_grid_folder(entry):
            continue
        _organize_one_grid(grid_dir, root_a, root_b, root_json, root_label)
        count += 1

    print(f"\n  [{channel}] 拢并完成，共 {count} 个网格 → {out_dir}")


def process_one_channel(output_dir: str, channel: str, step_base: int, step_total: int) -> int:
    """对单个通道执行：筛日期 → 重命名 → Excel → 拢并。返回下一步 step 编号。"""
    channel_dir = os.path.join(output_dir, channel)
    if not os.path.isdir(channel_dir):
        print(f"  [{channel}] 目录不存在，跳过")
        return step_base

    print(f"\n--- 通道 {channel}：{channel_dir} ---")

    print(f"[{step_base}/{step_total}] [{channel}] 筛选两个日期...")
    deleted = select_two_dates(channel_dir)
    print(f"  共删除 {deleted} 个多余日期文件夹")
    step_base += 1

    print(f"[{step_base}/{step_total}] [{channel}] 重命名图像...")
    n = rename_images(channel_dir, channel)
    print(f"  共重命名 {n} 张")
    step_base += 1

    print(f"[{step_base}/{step_total}] [{channel}] 生成统计 Excel...")
    generate_excel(channel_dir, channel)
    step_base += 1

    print(f"[{step_base}/{step_total}] [{channel}] GEOAI 拢并...")
    organize_geoai(channel_dir, output_dir, channel)
    step_base += 1

    return step_base


def main():
    print(f"输入目录：{INPUT_DIR}")
    print(f"输出目录：{OUTPUT_DIR}")
    print(f"通道：{', '.join(IMAGE_CHANNELS.keys())}")
    print("-" * 50)

    n_ch = len(IMAGE_CHANNELS)
    # 扫描 + 复制 + 每通道 4 步
    step_total = 2 + n_ch * 4

    step = 1
    print(f"[{step}/{step_total}] 扫描各通道图像...")
    records = collect_images(INPUT_DIR)
    by_ch = defaultdict(int)
    for rec in records:
        by_ch[rec["channel"]] += 1
    for ch in IMAGE_CHANNELS:
        print(f"  [{ch}] 待复制 {by_ch[ch]} 张")
    step += 1

    print(f"[{step}/{step_total}] 复制到 {OUTPUT_DIR}/{{VIS|WIDE|ZOOM}}/网格/日期/ ...")
    for ch in IMAGE_CHANNELS:
        os.makedirs(os.path.join(OUTPUT_DIR, ch), exist_ok=True)
    copy_images(records, OUTPUT_DIR)
    step += 1

    for channel in IMAGE_CHANNELS:
        step = process_one_channel(OUTPUT_DIR, channel, step, step_total)

    print("-" * 50)
    print("完成！")
    print(f"各通道 GEOAI 输出：{GEOAI_PREFIX}-VIS / -WIDE / -ZOOM（位于输出目录下）")


if __name__ == "__main__":
    main()
